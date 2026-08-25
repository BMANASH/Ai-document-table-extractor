import streamlit as st
import pandas as pd
import io
import json
import os
import re
import time
from PIL import Image
from pypdf import PdfReader
import google.generativeai as genai
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
import plotly.express as px
import plotly.graph_objects as go

# Page Configuration
st.set_page_config(
    page_title="SheetGen AI | Universal Document to Excel & Dashboard",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Fetch API Key silently from Streamlit Secrets or Environment
api_key = st.secrets.get("GEMINI_API_KEY") or os.environ.get("GEMINI_API_KEY", "")

# Modern Dark Theme, Universal Dark Mode Lock & Glassmorphic UI CSS
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700;800&family=Space+Grotesk:wght@600;700&display=swap');

html, body, [data-testid="stAppViewContainer"], .stApp {
    background: radial-gradient(circle at 50% 0%, #111827 0%, #080b11 75%, #05070a 100%) !important;
    background-color: #080b11 !important;
    color: #e2e8f0 !important;
    font-family: 'Plus Jakarta Sans', sans-serif !important;
}

header[data-testid="stHeader"], [data-testid="stHeader"] {
    background-color: rgba(8, 11, 17, 0.8) !important;
    backdrop-filter: blur(10px) !important;
    color: #e2e8f0 !important;
}

section[data-testid="stSidebar"] {
    min-width: 300px !important;
    max-width: 300px !important;
    width: 300px !important;
    background-color: #0b0f19 !important;
    border-right: 1px solid rgba(255, 255, 255, 0.08) !important;
}

[data-testid="stSidebarResizer"] {
    display: none !important;
    pointer-events: none !important;
}

h1, h2, h3, h4 {
    font-family: 'Space Grotesk', sans-serif !important;
    letter-spacing: -0.02em;
    font-weight: 700;
}

.header-wrapper {
    display: flex;
    align-items: center;
    gap: 16px;
    margin-top: 0.5rem;
    margin-bottom: 0.25rem;
}

.excel-title-text {
    background: linear-gradient(135deg, #22c55e 0%, #10b981 50%, #38bdf8 100%);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    font-size: 2.8rem;
    font-weight: 800;
    line-height: 1.15;
    font-family: 'Space Grotesk', sans-serif;
}

.sub-heading {
    color: #94a3b8;
    font-size: 1.05rem;
    font-weight: 400;
    margin-bottom: 1.75rem;
}

.steps-title {
    font-size: 0.85rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.08em;
    color: #38bdf8;
    margin-bottom: 0.75rem;
    display: flex;
    align-items: center;
    gap: 8px;
}

@keyframes floatCard1 {
    0% { transform: translateY(0px); }
    50% { transform: translateY(-7px); }
    100% { transform: translateY(0px); }
}
@keyframes floatCard2 {
    0% { transform: translateY(0px); }
    50% { transform: translateY(-5px); }
    100% { transform: translateY(0px); }
}
@keyframes floatCard3 {
    0% { transform: translateY(0px); }
    50% { transform: translateY(-8px); }
    100% { transform: translateY(0px); }
}

.feature-card {
    background: rgba(17, 24, 39, 0.7);
    border: 1px solid rgba(255, 255, 255, 0.08);
    border-radius: 14px;
    padding: 1.3rem;
    backdrop-filter: blur(12px);
    box-shadow: 0 8px 30px rgba(0, 0, 0, 0.35);
    transition: all 0.35s cubic-bezier(0.4, 0, 0.2, 1);
}

.card-anim-1 { animation: floatCard1 5.5s ease-in-out infinite; }
.card-anim-2 { animation: floatCard2 6.2s ease-in-out infinite 0.75s; }
.card-anim-3 { animation: floatCard3 5.8s ease-in-out infinite 1.5s; }

.feature-card:hover {
    transform: translateY(-9px) scale(1.02) !important;
    border-color: rgba(34, 197, 94, 0.45) !important;
    box-shadow: 0 16px 36px rgba(0, 0, 0, 0.5), 0 0 22px rgba(34, 197, 94, 0.25) !important;
}

.sidebar-item {
    background: rgba(17, 24, 39, 0.6);
    border: 1px solid rgba(255, 255, 255, 0.05);
    border-radius: 8px;
    padding: 10px 12px;
    margin-bottom: 8px;
}
.sidebar-title {
    font-size: 0.85rem;
    font-weight: 600;
    color: #e2e8f0;
    margin-bottom: 2px;
}
.sidebar-desc {
    font-size: 0.75rem;
    color: #94a3b8;
    line-height: 1.25;
}

.status-badge {
    display: inline-flex;
    align-items: center;
    gap: 6px;
    padding: 4px 12px;
    border-radius: 9999px;
    font-size: 0.75rem;
    font-weight: 600;
    background: rgba(34, 197, 94, 0.12);
    color: #4ade80;
    border: 1px solid rgba(34, 197, 94, 0.3);
    margin-bottom: 1rem;
}

div[data-testid="stFileUploader"] {
    background: radial-gradient(circle at 50% 50%, rgba(34, 197, 94, 0.08) 0%, rgba(17, 24, 39, 0.85) 100%) !important;
    border: 2px dashed rgba(34, 197, 94, 0.6) !important;
    border-radius: 16px !important;
    padding: 1.75rem 1.5rem !important;
    box-shadow: 0 8px 32px rgba(0, 0, 0, 0.45), 0 0 24px rgba(34, 197, 94, 0.15) !important;
    transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1) !important;
}

div[data-testid="stFileUploader"]:hover {
    border-color: #22c55e !important;
    background: radial-gradient(circle at 50% 50%, rgba(34, 197, 94, 0.14) 0%, rgba(17, 24, 39, 0.95) 100%) !important;
    box-shadow: 0 12px 40px rgba(0, 0, 0, 0.6), 0 0 35px rgba(34, 197, 94, 0.3) !important;
    transform: translateY(-2px);
}

div[data-testid="stFileUploader"] label {
    font-size: 1.12rem !important;
    font-weight: 700 !important;
    color: #ffffff !important;
    letter-spacing: -0.01em !important;
    margin-bottom: 0.6rem !important;
}

div[data-testid="stFileUploader"] section button {
    background: linear-gradient(135deg, #107C41 0%, #15803d 100%) !important;
    color: #ffffff !important;
    font-weight: 700 !important;
    border: 1px solid rgba(255, 255, 255, 0.2) !important;
    border-radius: 10px !important;
    padding: 0.55rem 1.35rem !important;
    box-shadow: 0 4px 16px rgba(16, 124, 65, 0.45) !important;
    transition: all 0.2s ease-in-out !important;
}

.kpi-stat-card {
    background: rgba(17, 24, 39, 0.75);
    border: 1px solid rgba(255, 255, 255, 0.08);
    border-radius: 14px;
    padding: 1.1rem 1.25rem;
    backdrop-filter: blur(12px);
    box-shadow: 0 6px 24px rgba(0, 0, 0, 0.3);
    margin-bottom: 12px;
}
.kpi-title {
    font-size: 0.8rem;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.05em;
    color: #94a3b8;
    margin-bottom: 4px;
}
.kpi-value {
    font-family: 'Space Grotesk', sans-serif;
    font-size: 1.8rem;
    font-weight: 800;
    color: #ffffff;
    line-height: 1.1;
}
.kpi-subtitle {
    font-size: 0.75rem;
    color: #38bdf8;
    margin-top: 4px;
}

[data-testid="stElementToolbar"] {
    opacity: 1 !important;
    visibility: visible !important;
    display: flex !important;
    gap: 6px !important;
    background: rgba(15, 23, 42, 0.95) !important;
    border: 1px solid rgba(34, 197, 94, 0.35) !important;
    border-radius: 10px !important;
    padding: 4px 8px !important;
    box-shadow: 0 6px 20px rgba(0, 0, 0, 0.5), 0 0 12px rgba(34, 197, 94, 0.2) !important;
    top: -46px !important;
    right: 0px !important;
}

[data-testid="stElementToolbar"] button {
    width: 34px !important;
    height: 34px !important;
    border-radius: 8px !important;
    background: rgba(255, 255, 255, 0.08) !important;
    color: #4ade80 !important;
    border: 1px solid rgba(255, 255, 255, 0.12) !important;
    transition: all 0.2s ease-in-out !important;
    display: flex !important;
    align-items: center !important;
    justify-content: center !important;
}

[data-testid="stElementToolbar"] button:hover {
    background: rgba(34, 197, 94, 0.25) !important;
    border-color: #22c55e !important;
    color: #ffffff !important;
    transform: scale(1.12) !important;
}

[data-testid="stElementToolbar"] svg {
    width: 18px !important;
    height: 18px !important;
    fill: currentColor !important;
}

.stDownloadButton > button {
    background: linear-gradient(135deg, #107C41 0%, #15803d 100%) !important;
    color: #ffffff !important;
    font-weight: 700 !important;
    font-size: 1.05rem !important;
    border: none !important;
    border-radius: 12px !important;
    padding: 0.85rem 2rem !important;
    box-shadow: 0 4px 25px rgba(16, 124, 65, 0.45) !important;
    transition: all 0.25s ease-in-out !important;
}
.stDownloadButton > button:hover {
    transform: translateY(-2px) !important;
    box-shadow: 0 6px 32px rgba(16, 124, 65, 0.75) !important;
}

.stButton > button[kind="primary"] {
    background: linear-gradient(135deg, #16a34a 0%, #15803d 100%) !important;
    color: #ffffff !important;
    font-weight: 600 !important;
    border: none !important;
    border-radius: 10px !important;
    padding: 0.65rem 1.75rem !important;
    box-shadow: 0 0 20px rgba(22, 163, 74, 0.4) !important;
}

@keyframes spinRadarRing {
    0% { transform: rotate(0deg); }
    100% { transform: rotate(360deg); }
}
@keyframes pulseGlassGlow {
    0% { box-shadow: 0 8px 32px rgba(0, 0, 0, 0.5), 0 0 18px rgba(34, 197, 94, 0.25); border-color: rgba(34, 197, 94, 0.4); }
    50% { box-shadow: 0 14px 44px rgba(0, 0, 0, 0.75), 0 0 36px rgba(34, 197, 94, 0.65); border-color: rgba(34, 197, 94, 0.95); }
    100% { box-shadow: 0 8px 32px rgba(0, 0, 0, 0.5), 0 0 18px rgba(34, 197, 94, 0.25); border-color: rgba(34, 197, 94, 0.4); }
}
@keyframes shimmerGlowText {
    0% { background-position: -200% 0; }
    100% { background-position: 200% 0; }
}

.glass-loading-card {
    background: rgba(15, 23, 42, 0.88) !important;
    backdrop-filter: blur(20px) !important;
    -webkit-backdrop-filter: blur(20px) !important;
    border: 1.5px solid rgba(34, 197, 94, 0.55) !important;
    border-radius: 18px !important;
    padding: 2.2rem 2.5rem !important;
    margin: 1.5rem 0 !important;
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: center;
    text-align: center;
    animation: pulseGlassGlow 2.5s infinite ease-in-out;
}

.spinner-radar-ring {
    width: 64px;
    height: 64px;
    border-radius: 50%;
    border: 3.5px solid rgba(255, 255, 255, 0.08);
    border-top: 3.5px solid #22c55e;
    border-right: 3.5px solid #38bdf8;
    animation: spinRadarRing 1.1s linear infinite;
    margin-bottom: 16px;
}

.glass-loading-title {
    font-family: 'Space Grotesk', sans-serif;
    font-size: 1.45rem;
    font-weight: 700;
    background: linear-gradient(90deg, #ffffff 0%, #4ade80 50%, #ffffff 100%);
    background-size: 200% auto;
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    animation: shimmerGlowText 3s linear infinite;
    margin-bottom: 8px;
}

.glass-loading-desc {
    font-size: 0.95rem;
    color: #94a3b8;
    max-width: 560px;
    line-height: 1.5;
    margin-bottom: 16px;
}

.status-pills-row {
    display: flex;
    flex-wrap: wrap;
    gap: 10px;
    justify-content: center;
}

.status-pill {
    font-size: 0.8rem;
    font-weight: 600;
    background: rgba(34, 197, 94, 0.14);
    color: #4ade80;
    border: 1px solid rgba(34, 197, 94, 0.35);
    padding: 5px 14px;
    border-radius: 9999px;
    display: inline-flex;
    align-items: center;
    gap: 6px;
}
</style>
""", unsafe_allow_html=True)

# Microsoft Excel Vector Badges
EXCEL_ICON_MAIN = '<svg width="48" height="48" viewBox="0 0 48 48" fill="none" xmlns="http://www.w3.org/2000/svg"><rect x="14" y="6" width="28" height="36" rx="4" fill="#107C41"/><rect x="23" y="13" width="6" height="4" fill="#ffffff" fill-opacity="0.85"/><rect x="31" y="13" width="6" height="4" fill="#ffffff" fill-opacity="0.85"/><rect x="23" y="19" width="6" height="4" fill="#ffffff" fill-opacity="0.85"/><rect x="31" y="19" width="6" height="4" fill="#ffffff" fill-opacity="0.85"/><rect x="23" y="25" width="6" height="4" fill="#ffffff" fill-opacity="0.85"/><rect x="31" y="25" width="6" height="4" fill="#ffffff" fill-opacity="0.85"/><rect x="23" y="31" width="6" height="4" fill="#ffffff" fill-opacity="0.85"/><rect x="31" y="31" width="6" height="4" fill="#ffffff" fill-opacity="0.85"/><rect x="6" y="9" width="22" height="30" rx="3" fill="#185C37" style="filter: drop-shadow(2px 4px 6px rgba(0,0,0,0.5));"/><path d="M11.5 30L15.3 24L11.8 18H15.2L17 21.5L18.8 18H22.2L18.7 24L22.5 30H19.1L17 26.2L14.9 30H11.5Z" fill="white"/></svg>'

EXCEL_ICON_SIDEBAR = '<svg width="28" height="28" viewBox="0 0 48 48" fill="none" xmlns="http://www.w3.org/2000/svg" style="vertical-align: middle;"><rect x="14" y="6" width="28" height="36" rx="4" fill="#107C41"/><rect x="23" y="13" width="6" height="4" fill="#ffffff" fill-opacity="0.85"/><rect x="31" y="13" width="6" height="4" fill="#ffffff" fill-opacity="0.85"/><rect x="23" y="19" width="6" height="4" fill="#ffffff" fill-opacity="0.85"/><rect x="31" y="19" width="6" height="4" fill="#ffffff" fill-opacity="0.85"/><rect x="23" y="25" width="6" height="4" fill="#ffffff" fill-opacity="0.85"/><rect x="31" y="25" width="6" height="4" fill="#ffffff" fill-opacity="0.85"/><rect x="23" y="31" width="6" height="4" fill="#ffffff" fill-opacity="0.85"/><rect x="31" y="31" width="6" height="4" fill="#ffffff" fill-opacity="0.85"/><rect x="6" y="9" width="22" height="30" rx="3" fill="#185C37"/><path d="M11.5 30L15.3 24L11.8 18H15.2L17 21.5L18.8 18H22.2L18.7 24L22.5 30H19.1L17 26.2L14.9 30H11.5Z" fill="white"/></svg>'

# Static Sidebar Configuration
with st.sidebar:
    st.markdown(f'<div style="display:flex; align-items:center; gap:10px; margin-bottom: 6px;">{EXCEL_ICON_SIDEBAR}<span style="font-size:1.35rem; font-weight:700; color:#ffffff; font-family:\'Space Grotesk\',sans-serif;">SheetGen AI</span></div>', unsafe_allow_html=True)
    st.caption("Universal Tabular Extraction & AI Dashboard Suite")
    st.markdown("<hr style='border:none; border-top:1px solid rgba(255,255,255,0.08); margin:12px 0;'>", unsafe_allow_html=True)
    
    st.markdown("<div style='font-size:0.8rem; font-weight:700; text-transform:uppercase; letter-spacing:0.05em; color:#94a3b8; margin-bottom:10px;'>Core Capabilities</div>", unsafe_allow_html=True)
    
    sidebar_items_html = (
        '<div class="sidebar-item"><div class="sidebar-title">⚡ Universal OCR Vision</div>'
        '<div class="sidebar-desc">Extracts tables from invoices, registers, ledgers, and receipts.</div></div>'
        '<div class="sidebar-item"><div class="sidebar-title">📊 Executive Dashboard Engine</div>'
        '<div class="sidebar-desc">Instant KPI metrics and interactive dark-glass visual charts.</div></div>'
        '<div class="sidebar-item"><div class="sidebar-title">💬 Talk with AI (Copilot)</div>'
        '<div class="sidebar-desc">Describe any custom visual chart and see it render instantly.</div></div>'
        '<div class="sidebar-item"><div class="sidebar-title">✏️ Visual In-Browser Grid</div>'
        '<div class="sidebar-desc">Enlarged toolbar to add rows, search, hide columns & edit data.</div></div>'
        '<div class="sidebar-item"><div class="sidebar-title">📥 Native .xlsx Generator</div>'
        '<div class="sidebar-desc">Multi-sheet Excel workbook with formatted master tab.</div></div>'
    )
    st.markdown(sidebar_items_html, unsafe_allow_html=True)
    st.markdown("<hr style='border:none; border-top:1px solid rgba(255,255,255,0.08); margin:12px 0;'>", unsafe_allow_html=True)
    st.markdown("🟢 **System Status:** Ready")

# Hero Header
st.markdown('<div class="status-badge">⚡ Instant Document to Spreadsheet & Dashboard</div>', unsafe_allow_html=True)
main_header_html = f'<div class="header-wrapper">{EXCEL_ICON_MAIN}<div class="excel-title-text">SheetGen AI</div></div>'
st.markdown(main_header_html, unsafe_allow_html=True)
st.markdown('<div class="sub-heading">Upload single or batch images & PDFs → Convert to styled Excel workbooks & interactive AI visual dashboards.</div>', unsafe_allow_html=True)

# 3 Step Visual Workflow Guide Header
st.markdown('<div class="steps-title">📋 Steps to Convert, Visualize & Download</div>', unsafe_allow_html=True)

col_card1, col_card2, col_card3 = st.columns(3)
with col_card1:
    card1_html = (
        '<div class="feature-card card-anim-1">'
        '<div style="font-size: 1.15rem; font-weight:700; color:#4ade80; margin-bottom:4px;">1. Upload File(s)</div>'
        '<div style="font-size:0.85rem; color:#94a3b8; line-height:1.4;">Drop photos or PDFs containing tabular registers below.</div>'
        '</div>'
    )
    st.markdown(card1_html, unsafe_allow_html=True)

with col_card2:
    card2_html = (
        '<div class="feature-card card-anim-2">'
        '<div style="font-size: 1.15rem; font-weight:700; color:#38bdf8; margin-bottom:4px;">2. AI Extracts & Cleans</div>'
        '<div style="font-size:0.85rem; color:#94a3b8; line-height:1.4;">Extracts rows, columns, and generates interactive dashboards.</div>'
        '</div>'
    )
    st.markdown(card2_html, unsafe_allow_html=True)

with col_card3:
    card3_html = (
        '<div class="feature-card card-anim-3">'
        '<div style="font-size: 1.15rem; font-weight:700; color:#a78bfa; margin-bottom:4px;">3. Talk with AI & Download</div>'
        '<div style="font-size:0.85rem; color:#94a3b8; line-height:1.4;">Create custom charts on demand and download the styled .xlsx.</div>'
        '</div>'
    )
    st.markdown(card3_html, unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

# Fast In-Memory Image Resizer (Optimized for Sub-10s Vision Processing)
def prepare_image(raw_bytes):
    img = Image.open(io.BytesIO(raw_bytes))
    if img.mode != "RGB":
        img = img.convert("RGB")
    max_dim = 1280
    if max(img.size) > max_dim:
        img.thumbnail((max_dim, max_dim), Image.Resampling.LANCZOS)
    out = io.BytesIO()
    img.save(out, format="JPEG", quality=80, optimize=True)
    out.seek(0)
    return Image.open(out)

# Universal Column Normalization
def normalize_column_header(col_name):
    c = str(col_name).strip()
    c_clean = re.sub(r'[^\w\s]', '', c).lower()
    if c_clean in ['no', 'sr no', 'srno', 'sl no', 'slno', 's no', 'sno', 'serial no', 'serial number']:
        return 'SL. NO.'
    return c.upper()

# Dynamic Dataset Statistical Profiler for AI Context
def profile_dataset_metrics(df):
    summary = {
        "total_records": len(df),
        "columns": list(df.columns),
        "column_profiles": {}
    }
    for c in df.columns:
        clean_s = df[c].astype(str).str.strip().str.upper().replace(["", "NAN", "NONE", "NULL"], pd.NA).dropna()
        num_s = pd.to_numeric(clean_s.str.replace(r'[^\d.]', '', regex=True), errors='coerce').dropna()
        if len(num_s) > len(clean_s) * 0.7 and c not in ['SL. NO.', 'NO.', 'PHONE', 'PHONE NUMBER']:
            summary["column_profiles"][c] = {
                "type": "numeric",
                "sum": float(num_s.sum()),
                "mean": float(num_s.mean()),
                "min": float(num_s.min()),
                "max": float(num_s.max())
            }
        else:
            top_counts = clean_s.value_counts().head(5).to_dict()
            summary["column_profiles"][c] = {
                "type": "categorical",
                "unique_count": int(clean_s.nunique()),
                "top_5_frequencies": top_counts
            }
    return summary

# =========================================================================
# EXCEL GENERATOR 1: BASE DATA ONLY (.XLSX)
# =========================================================================
def generate_base_excel_workbook(sheets_map):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
    regular_font = Font(name="Calibri", size=10, color="1F2937")
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'), right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'), bottom=Side(style='thin', color='E5E7EB')
    )
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    
    for title, df in sheets_map.items():
        ws = wb.create_sheet(title=title[:31])
        ws.views.sheetView[0].showGridLines = True
        
        headers = list(df.columns)
        ws.append(headers)
        ws.row_dimensions[1].height = 26
        
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
            
        for r_idx, row in enumerate(df.itertuples(index=False), start=2):
            ws.row_dimensions[r_idx].height = 20
            is_even = (r_idx % 2 == 0)
            for col_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=col_idx)
                val_clean = "" if pd.isna(val) or str(val).strip().lower() in ["none", "nan", "null"] else str(val).strip()
                cell.value = val_clean
                cell.font = regular_font
                cell.border = thin_border
                
                if (val_clean.isdigit() and len(val_clean) < 5) or headers[col_idx-1] == 'SL. NO.':
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    
                if is_even:
                    cell.fill = zebra_fill
                    
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_s = str(cell.value or "")
                if len(val_s) > max_len:
                    max_len = len(val_s)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 14)
            
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# =========================================================================
# EXCEL GENERATOR 2: EXECUTIVE DATA DASHBOARD WITH LIVE FORMULAS (.XLSX)
# =========================================================================
def generate_smart_dashboard_excel_workbook(sheets_map, master_df):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    BRAND_GREEN = "107C41"
    CARD_BG = "F8FAFC"
    CARD_BORDER_CLR = "CBD5E1"
    TEXT_DARK = "0F172A"
    TEXT_MUTED = "64748B"

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=BRAND_GREEN, end_color=BRAND_GREEN, fill_type="solid")
    card_fill = PatternFill(start_color=CARD_BG, end_color=CARD_BG, fill_type="solid")
    regular_font = Font(name="Calibri", size=10, color=TEXT_DARK)
    
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'), right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'), bottom=Side(style='thin', color='E5E7EB')
    )
    thin_card_border = Border(
        left=Side(style='thin', color=CARD_BORDER_CLR), right=Side(style='thin', color=CARD_BORDER_CLR),
        top=Side(style='thin', color=CARD_BORDER_CLR), bottom=Side(style='thin', color=CARD_BORDER_CLR)
    )
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    
    data_sheet_name = "Master Combined Records" if len(sheets_map) > 1 else list(sheets_map.keys())[0]
    
    # 1. Populate Raw Data Sheets
    for title, df in sheets_map.items():
        ws = wb.create_sheet(title=title[:31])
        ws.views.sheetView[0].showGridLines = True
        headers = list(df.columns)
        ws.append(headers)
        ws.row_dimensions[1].height = 26
        
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
            
        for r_idx, row in enumerate(df.itertuples(index=False), start=2):
            ws.row_dimensions[r_idx].height = 20
            is_even = (r_idx % 2 == 0)
            for col_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=col_idx)
                val_clean = "" if pd.isna(val) or str(val).strip().lower() in ["none", "nan", "null"] else str(val).strip()
                cell.value = val_clean
                cell.font = regular_font
                cell.border = thin_border
                if (val_clean.isdigit() and len(val_clean) < 5) or headers[col_idx-1] == 'SL. NO.':
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                if is_even:
                    cell.fill = zebra_fill
                    
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_s = str(cell.value or "")
                if len(val_s) > max_len:
                    max_len = len(val_s)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 14)
            
    # 2. Build Executive Formula Dashboard Sheet at Index 0
    ws_dash = wb.create_sheet(title="Executive Dashboard", index=0)
    ws_dash.views.sheetView[0].showGridLines = True
    
    # Title Banner
    ws_dash.merge_cells("B2:R2")
    t_cell = ws_dash["B2"]
    t_cell.value = "   📊 EXECUTIVE DATA INTELLIGENCE DASHBOARD"
    t_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    t_cell.fill = header_fill
    t_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws_dash.row_dimensions[2].height = 36
    
    cat_candidates = [c for c in master_df.columns if c not in ['SL. NO.', 'EMPLOYEE NAME', 'PHONE', 'PHONE NUMBER', 'NO.']]
    cat1 = cat_candidates[0] if len(cat_candidates) > 0 else master_df.columns[0]
    cat2 = cat_candidates[1] if len(cat_candidates) > 1 else cat1
    
    clean_series1 = master_df[cat1].astype(str).str.strip().str.upper().replace("", pd.NA).dropna()
    unique_cnt1 = clean_series1.nunique()
    
    clean_series2 = master_df[cat2].astype(str).str.strip().str.upper().replace("", pd.NA).dropna()
    top_val2 = clean_series2.mode().iloc[0] if not clean_series2.empty else "N/A"
    
    # 3 Corporate KPI Cards
    ws_dash.merge_cells("B4:D4")
    ws_dash.merge_cells("B5:D5")
    ws_dash["B4"].value = "TOTAL RECORDS PARSED"
    ws_dash["B4"].font = Font(name="Calibri", size=9, bold=True, color=TEXT_MUTED)
    ws_dash["B5"].value = f"=COUNTA('{data_sheet_name}'!B2:B{len(master_df)+1})"
    ws_dash["B5"].font = Font(name="Calibri", size=20, bold=True, color=BRAND_GREEN)
    for r in range(4, 6):
        for c in range(2, 5):
            cell = ws_dash.cell(row=r, column=c)
            cell.fill = card_fill
            cell.border = thin_card_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws_dash.merge_cells("F4:H4")
    ws_dash.merge_cells("F5:H5")
    ws_dash["F4"].value = f"UNIQUE {cat1[:15]}"
    ws_dash["F4"].font = Font(name="Calibri", size=9, bold=True, color=TEXT_MUTED)
    ws_dash["F5"].value = unique_cnt1
    ws_dash["F5"].font = Font(name="Calibri", size=20, bold=True, color="0284C7")
    for r in range(4, 6):
        for c in range(6, 9):
            cell = ws_dash.cell(row=r, column=c)
            cell.fill = card_fill
            cell.border = thin_card_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws_dash.merge_cells("J4:L4")
    ws_dash.merge_cells("J5:L5")
    ws_dash["J4"].value = f"DOMINANT {cat2[:14]}"
    ws_dash["J4"].font = Font(name="Calibri", size=9, bold=True, color=TEXT_MUTED)
    ws_dash["J5"].value = str(top_val2)[:18]
    ws_dash["J5"].font = Font(name="Calibri", size=18, bold=True, color=BRAND_GREEN)
    for r in range(4, 6):
        for c in range(10, 13):
            cell = ws_dash.cell(row=r, column=c)
            cell.fill = card_fill
            cell.border = thin_card_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws_dash.row_dimensions[4].height = 18
    ws_dash.row_dimensions[5].height = 30
    
    # 4. Summary Table 1 & Native Bar Chart
    c1_idx = master_df.columns.get_loc(cat1) + 1
    c1_let = get_column_letter(c1_idx)
    
    ws_dash["B8"].value = cat1
    ws_dash["C8"].value = "RECORD COUNT"
    ws_dash["B8"].font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    ws_dash["B8"].fill = header_fill
    ws_dash["C8"].font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    ws_dash["C8"].fill = header_fill
    
    unique_cats1 = clean_series1.value_counts().head(8).index.tolist()
    for idx, val in enumerate(unique_cats1, start=9):
        ws_dash[f"B{idx}"].value = str(val)
        ws_dash[f"C{idx}"].value = f"=COUNTIF('{data_sheet_name}'!${c1_let}$2:${c1_let}${len(master_df)+1}, B{idx})"
        ws_dash[f"B{idx}"].font = regular_font
        ws_dash[f"C{idx}"].font = regular_font
        ws_dash[f"B{idx}"].border = thin_card_border
        ws_dash[f"C{idx}"].border = thin_card_border
        ws_dash[f"C{idx}"].alignment = Alignment(horizontal="center", vertical="center")
        
    tot_row1 = 9 + len(unique_cats1)
    ws_dash[f"B{tot_row1}"].value = "TOTAL"
    ws_dash[f"B{tot_row1}"].font = Font(name="Calibri", size=10, bold=True, color=TEXT_DARK)
    ws_dash[f"C{tot_row1}"].value = f"=SUM(C9:C{tot_row1-1})"
    ws_dash[f"C{tot_row1}"].font = Font(name="Calibri", size=10, bold=True, color=TEXT_DARK)
    ws_dash[f"B{tot_row1}"].border = thin_card_border
    ws_dash[f"C{tot_row1}"].border = thin_card_border
    ws_dash[f"C{tot_row1}"].alignment = Alignment(horizontal="center", vertical="center")

    if len(unique_cats1) > 0:
        bar = BarChart()
        bar.type = "col"
        bar.style = 10
        bar.title = f"Distribution by {cat1}"
        bar.y_axis.title = "Count"
        bar.x_axis.title = cat1
        bar.varyColors = True
        bar.legend = None
        bar.gapWidth = 80
        
        bar.dataLabels = DataLabelList()
        bar.dataLabels.showVal = True
        bar.dataLabels.showCatName = False
        bar.dataLabels.showSerName = False
        bar.dataLabels.showPercent = False
        bar.dataLabels.showLegendKey = False
        
        data_ref1 = Reference(ws_dash, min_col=3, min_row=8, max_row=8 + len(unique_cats1))
        cats_ref1 = Reference(ws_dash, min_col=2, min_row=9, max_row=8 + len(unique_cats1))
        bar.add_data(data_ref1, titles_from_data=True)
        bar.set_categories(cats_ref1)
        bar.height = 11
        bar.width = 16
        ws_dash.add_chart(bar, "E8")

    # 5. Summary Table 2 & Native Doughnut Chart
    start_s = tot_row1 + 3
    c2_idx = master_df.columns.get_loc(cat2) + 1
    c2_let = get_column_letter(c2_idx)
    
    ws_dash[f"B{start_s}"].value = cat2
    ws_dash[f"C{start_s}"].value = "RECORD COUNT"
    ws_dash[f"B{start_s}"].font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    ws_dash[f"B{start_s}"].fill = header_fill
    ws_dash[f"C{start_s}"].font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    ws_dash[f"C{start_s}"].fill = header_fill
    
    unique_cats2 = clean_series2.value_counts().head(6).index.tolist()
    for idx, val in enumerate(unique_cats2, start=start_s + 1):
        ws_dash[f"B{idx}"].value = str(val)
        ws_dash[f"C{idx}"].value = f"=COUNTIF('{data_sheet_name}'!${c2_let}$2:${c2_let}${len(master_df)+1}, B{idx})"
        ws_dash[f"B{idx}"].font = regular_font
        ws_dash[f"C{idx}"].font = regular_font
        ws_dash[f"B{idx}"].border = thin_card_border
        ws_dash[f"C{idx}"].border = thin_card_border
        ws_dash[f"C{idx}"].alignment = Alignment(horizontal="center", vertical="center")

    tot_row2 = start_s + 1 + len(unique_cats2)
    ws_dash[f"B{tot_row2}"].value = "TOTAL"
    ws_dash[f"B{tot_row2}"].font = Font(name="Calibri", size=10, bold=True, color=TEXT_DARK)
    ws_dash[f"C{tot_row2}"].value = f"=SUM(C{start_s+1}:C{tot_row2-1})"
    ws_dash[f"C{tot_row2}"].font = Font(name="Calibri", size=10, bold=True, color=TEXT_DARK)
    ws_dash[f"B{tot_row2}"].border = thin_card_border
    ws_dash[f"C{tot_row2}"].border = thin_card_border
    ws_dash[f"C{tot_row2}"].alignment = Alignment(horizontal="center", vertical="center")

    if len(unique_cats2) > 0:
        donut = DoughnutChart()
        donut.title = f"Proportion of {cat2}"
        donut.holeSize = 55
        
        donut.dataLabels = DataLabelList()
        donut.dataLabels.showPercent = True
        donut.dataLabels.showVal = False
        donut.dataLabels.showCatName = False
        donut.dataLabels.showSerName = False
        donut.dataLabels.showLegendKey = False
        if donut.legend:
            donut.legend.legendPos = "r"
            
        data_ref2 = Reference(ws_dash, min_col=3, min_row=start_s, max_row=start_s + len(unique_cats2))
        cats_ref2 = Reference(ws_dash, min_col=2, min_row=start_s + 1, max_row=start_s + len(unique_cats2))
        donut.add_data(data_ref2, titles_from_data=True)
        donut.set_categories(cats_ref2)
        donut.height = 11
        donut.width = 15
        ws_dash.add_chart(donut, "L8")

    for col in ws_dash.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_s = str(cell.value or "")
            if len(val_s) > max_len and not val_s.startswith("="):
                max_len = len(val_s)
        ws_dash.column_dimensions[col_letter].width = max(max_len + 4, 16)
        
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# =========================================================================
# EXCEL GENERATOR 3: DATA + AI CUSTOM COPILOT DASHBOARD (.XLSX)
# =========================================================================
def generate_ai_copilot_excel_workbook(sheets_map, master_df, ai_spec):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    BRAND_GREEN = "107C41"
    CARD_BG = "F8FAFC"
    CARD_BORDER_CLR = "CBD5E1"
    TEXT_DARK = "0F172A"
    TEXT_MUTED = "64748B"

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=BRAND_GREEN, end_color=BRAND_GREEN, fill_type="solid")
    card_fill = PatternFill(start_color=CARD_BG, end_color=CARD_BG, fill_type="solid")
    regular_font = Font(name="Calibri", size=10, color=TEXT_DARK)
    
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'), right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'), bottom=Side(style='thin', color='E5E7EB')
    )
    thin_card_border = Border(
        left=Side(style='thin', color=CARD_BORDER_CLR), right=Side(style='thin', color=CARD_BORDER_CLR),
        top=Side(style='thin', color=CARD_BORDER_CLR), bottom=Side(style='thin', color=CARD_BORDER_CLR)
    )
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    
    data_sheet_name = "Master Combined Records" if len(sheets_map) > 1 else list(sheets_map.keys())[0]
    
    for title, df in sheets_map.items():
        ws = wb.create_sheet(title=title[:31])
        ws.views.sheetView[0].showGridLines = True
        headers = list(df.columns)
        ws.append(headers)
        ws.row_dimensions[1].height = 26
        
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
            
        for r_idx, row in enumerate(df.itertuples(index=False), start=2):
            ws.row_dimensions[r_idx].height = 20
            is_even = (r_idx % 2 == 0)
            for col_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=col_idx)
                val_clean = "" if pd.isna(val) or str(val).strip().lower() in ["none", "nan", "null"] else str(val).strip()
                cell.value = val_clean
                cell.font = regular_font
                cell.border = thin_border
                if (val_clean.isdigit() and len(val_clean) < 5) or headers[col_idx-1] == 'SL. NO.':
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                if is_even:
                    cell.fill = zebra_fill
                    
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_s = str(cell.value or "")
                if len(val_s) > max_len:
                    max_len = len(val_s)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 14)
            
    # Write Custom AI Dashboard Sheet at Index 0
    ws_ai = wb.create_sheet(title="Custom AI Dashboard", index=0)
    ws_ai.views.sheetView[0].showGridLines = True
    
    ws_ai.merge_cells("B2:R2")
    t_cell = ws_ai["B2"]
    t_cell.value = "   🤖 AI COPILOT CUSTOM VISUAL DASHBOARD"
    t_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    t_cell.fill = PatternFill(start_color="0284C7", end_color="0284C7", fill_type="solid")
    t_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws_ai.row_dimensions[2].height = 36
    
    # Custom KPIs with live formula on Card 1
    kpis = ai_spec.get("kpi_cards", [])
    for idx, k in enumerate(kpis[:3]):
        c_start = 2 + idx * 4
        c_end = c_start + 2
        col_s_let = get_column_letter(c_start)
        col_e_let = get_column_letter(c_end)
        
        ws_ai.merge_cells(f"{col_s_let}4:{col_e_let}4")
        ws_ai.merge_cells(f"{col_s_let}5:{col_e_let}5")
        ws_ai[f"{col_s_let}4"].value = str(k.get("label", "Metric")).upper()
        ws_ai[f"{col_s_let}4"].font = Font(name="Calibri", size=9, bold=True, color=TEXT_MUTED)
        
        if idx == 0 and "TOTAL" in str(k.get("label", "")).upper():
            ws_ai[f"{col_s_let}5"].value = f"=COUNTA('{data_sheet_name}'!B2:B{len(master_df)+1})"
        else:
            ws_ai[f"{col_s_let}5"].value = str(k.get("value", "-"))
            
        ws_ai[f"{col_s_let}5"].font = Font(name="Calibri", size=18, bold=True, color="0284C7")
        
        for r in range(4, 6):
            for c in range(c_start, c_end + 1):
                cell = ws_ai.cell(row=r, column=c)
                cell.fill = card_fill
                cell.border = thin_card_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
    ws_ai.row_dimensions[4].height = 18
    ws_ai.row_dimensions[5].height = 30
    
    charts = ai_spec.get("charts", [])
    start_row = 8
    chart_positions = ["E8", "L8"]
    
    for c_idx, c_info in enumerate(charts[:2]):
        g_col = c_info.get("group_by_col")
        if g_col not in master_df.columns:
            g_col = master_df.columns[0]
            
        col_idx_data = master_df.columns.get_loc(g_col) + 1
        col_let_data = get_column_letter(col_idx_data)
        
        ws_ai[f"B{start_row}"].value = g_col
        ws_ai[f"C{start_row}"].value = "CALCULATED COUNT"
        ws_ai[f"B{start_row}"].font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        ws_ai[f"B{start_row}"].fill = header_fill
        ws_ai[f"C{start_row}"].font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        ws_ai[f"C{start_row}"].fill = header_fill
        
        clean_s = master_df[g_col].astype(str).str.strip().str.upper().replace("", pd.NA).dropna()
        cats = clean_s.value_counts().head(8).index.tolist()
        
        for r_offset, c_val in enumerate(cats, start=1):
            curr_row = start_row + r_offset
            ws_ai[f"B{curr_row}"].value = str(c_val)
            ws_ai[f"C{curr_row}"].value = f"=COUNTIF('{data_sheet_name}'!${col_let_data}$2:${col_let_data}${len(master_df)+1}, B{curr_row})"
            ws_ai[f"B{curr_row}"].font = regular_font
            ws_ai[f"C{curr_row}"].font = regular_font
            ws_ai[f"B{curr_row}"].border = thin_card_border
            ws_ai[f"C{curr_row}"].border = thin_card_border
            ws_ai[f"C{curr_row}"].alignment = Alignment(horizontal="center", vertical="center")
            
        tot_row = start_row + 1 + len(cats)
        ws_ai[f"B{tot_row}"].value = "TOTAL"
        ws_ai[f"B{tot_row}"].font = Font(name="Calibri", size=10, bold=True, color=TEXT_DARK)
        ws_ai[f"C{tot_row}"].value = f"=SUM(C{start_row+1}:C{tot_row-1})"
        ws_ai[f"C{tot_row}"].font = Font(name="Calibri", size=10, bold=True, color=TEXT_DARK)
        ws_ai[f"B{tot_row}"].border = thin_card_border
        ws_ai[f"C{tot_row}"].border = thin_card_border
        ws_ai[f"C{tot_row}"].alignment = Alignment(horizontal="center", vertical="center")

        if len(cats) > 0:
            c_type = c_info.get("chart_type", "bar").lower()
            if c_type in ["pie", "donut"]:
                ai_chart = DoughnutChart()
                ai_chart.holeSize = 55
                ai_chart.dataLabels = DataLabelList()
                ai_chart.dataLabels.showPercent = True
                ai_chart.dataLabels.showVal = False
                ai_chart.dataLabels.showCatName = False
                ai_chart.dataLabels.showSerName = False
                ai_chart.dataLabels.showLegendKey = False
                if ai_chart.legend:
                    ai_chart.legend.legendPos = "r"
            else:
                ai_chart = BarChart()
                ai_chart.type = "col"
                ai_chart.legend = None
                ai_chart.varyColors = True
                ai_chart.gapWidth = 80
                ai_chart.y_axis.title = "Count"
                ai_chart.x_axis.title = g_col
                ai_chart.dataLabels = DataLabelList()
                ai_chart.dataLabels.showVal = True
                ai_chart.dataLabels.showCatName = False
                ai_chart.dataLabels.showSerName = False
                ai_chart.dataLabels.showPercent = False
                ai_chart.dataLabels.showLegendKey = False
                
            ai_chart.title = c_info.get("title", f"Summary of {g_col}")
            ai_chart.style = 10
            data_ref_ai = Reference(ws_ai, min_col=3, min_row=start_row, max_row=start_row + len(cats))
            cats_ref_ai = Reference(ws_ai, min_col=2, min_row=start_row + 1, max_row=start_row + len(cats))
            ai_chart.add_data(data_ref_ai, titles_from_data=True)
            ai_chart.set_categories(cats_ref_ai)
            ai_chart.height = 11
            ai_chart.width = 16 if c_idx == 0 else 15
            ws_ai.add_chart(ai_chart, chart_positions[c_idx])
            
        start_row = tot_row + 3
        
    for col in ws_ai.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_s = str(cell.value or "")
            if len(val_s) > max_len and not val_s.startswith("="):
                max_len = len(val_s)
        ws_ai.column_dimensions[col_letter].width = max(max_len + 4, 16)
        
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# =========================================================================
# HIGH-SPEED EXTRACTION ENGINE (8-15s TARGET)
# =========================================================================
def execute_extraction_cascade(files_data, key_str):
    genai.configure(api_key=key_str)
    
    prompt = """
    You are an expert Data Engineer and OCR Analyst.
    Extract all tabular information from the uploaded file(s) with high fidelity and precision:

    1. TABLE EXTRACTION (Domain Agnostic):
       - Accurately identify all distinct tables across all uploaded pages and files (invoices, ledgers, registers, bills, spreadsheets).
       - Accurately preserve and transcribe the authentic column headers present in the document.
       - Clean values: If a cell is blank or unreadable, return an empty string "". Do NOT write literal "None" or "NaN".
       - Assign an intuitive title for each table/sheet (e.g. "Invoice Breakdown", "Page 1 - Ledger", "Attendance Roster").

    2. EXECUTIVE SUMMARY:
       - Provide a concise, structured business summary of the extracted data in Markdown format:
         * 📌 **Document Type & Scope**: What type of document this is.
         * 📊 **Key Metrics & Statistics**: Total record count, key numeric totals or categories.
         * 🔍 **Observations**: Noteworthy trends, missing fields, or exceptions.

    Return output strictly as valid JSON matching this schema:
    {
      "analysis": "Structured summary text in Markdown with bold headers and bullet points.",
      "tables": [
        {
          "table_name": "Table Name",
          "headers": ["Header 1", "Header 2", "Header 3"],
          "rows": [
            ["Row1 Col1", "Row1 Col2", "Row1 Col3"]
          ]
        }
      ]
    }
    """
    
    contents = []
    for file_bytes, mime_type in files_data:
        if "image" in mime_type:
            contents.append(prepare_image(file_bytes))
        elif "pdf" in mime_type:
            reader = PdfReader(io.BytesIO(file_bytes))
            pdf_text = "\n".join([p.extract_text() or "" for p in reader.pages])
            contents.append(f"PDF Content:\n{pdf_text}")
            
    contents.append(prompt)

    # Primary high-speed vision models
    fast_models = ["gemini-2.5-flash", "gemini-2.0-flash", "gemini-1.5-flash"]
    last_err = None
    
    for model_name in fast_models:
        try:
            model = genai.GenerativeModel(model_name)
            response = model.generate_content(
                contents,
                generation_config={"response_mime_type": "application/json"}
            )
            if response and response.text:
                raw_text = response.text.strip()
                if "```json" in raw_text:
                    raw_text = raw_text.split("```json")[1].split("```")[0].strip()
                elif "```" in raw_text:
                    raw_text = raw_text.split("```")[1].split("```")[0].strip()
                return raw_text, model_name
        except Exception as err:
            last_err = err
            continue

    raise Exception(f"Extraction failed. Last Error: {last_err}")

# High-Speed AI Custom Dashboard Copilot Function
def ask_ai_for_dashboard_spec(df, user_instruction, key_str):
    genai.configure(api_key=key_str)
    metrics_context = profile_dataset_metrics(df)
    
    prompt = f"""
    You are an expert Data Visualizer and Business Intelligence Architect.
    Here is the complete statistical profile of the dataset:
    {json.dumps(metrics_context, indent=2)}
    
    The user requested this dashboard / visualization:
    "{user_instruction}"
    
    IMPORTANT RULES:
    1. Base all KPI metric cards on the true statistical profile (e.g. total_records is {metrics_context['total_records']}).
    2. Suggest 1 or 2 high-impact charts selecting exact column names from: {list(df.columns)}.
    3. Write accurate business insights summarizing the findings.
    
    Return strictly valid JSON matching this schema:
    {{
      "kpi_cards": [
        {{"label": "Metric Name", "value": "Exact Calculated Stat", "subtitle": "Context or note"}}
      ],
      "charts": [
        {{
          "chart_type": "bar", // Choose from: "bar", "pie", "donut", "line", "treemap"
          "title": "Clean Chart Title",
          "group_by_col": "Exact Column Name",
          "metric_col": null,
          "aggregation": "count",
          "top_n": 8
        }}
      ],
      "ai_insights": "2-3 lines explaining true insights based on data."
    }}
    """
    
    fast_models = ["gemini-2.5-flash", "gemini-2.0-flash", "gemini-1.5-flash"]
    for model_name in fast_models:
        try:
            model = genai.GenerativeModel(model_name)
            res = model.generate_content(
                prompt, 
                generation_config={"response_mime_type": "application/json"}
            )
            if res and res.text:
                t = res.text.strip()
                if "```json" in t:
                    t = t.split("```json")[1].split("```")[0].strip()
                elif "```" in t:
                    t = t.split("```")[1].split("```")[0].strip()
                return json.loads(t)
        except Exception:
            continue
    return None

# Render Dynamic Plotly Chart in Dark Glass Theme
def render_plotly_chart(df, chart_spec):
    chart_type = chart_spec.get("chart_type", "bar").lower()
    title = chart_spec.get("title", "Chart")
    group_col = chart_spec.get("group_by_col")
    metric_col = chart_spec.get("metric_col")
    agg = chart_spec.get("aggregation", "count").lower()
    top_n = chart_spec.get("top_n", 8)
    
    if group_col not in df.columns:
        text_cols = [c for c in df.columns if c not in ['SL. NO.', 'NO.']]
        group_col = text_cols[0] if text_cols else df.columns[0]
        
    plot_df = df[df[group_col].astype(str).str.strip() != ""].copy()
    
    if metric_col and metric_col in plot_df.columns:
        plot_df[metric_col] = pd.to_numeric(plot_df[metric_col].astype(str).str.replace(r'[^\d.]', '', regex=True), errors='coerce').fillna(0)
        if agg == "sum":
            agg_df = plot_df.groupby(group_col)[metric_col].sum().reset_index()
        elif agg == "mean":
            agg_df = plot_df.groupby(group_col)[metric_col].mean().reset_index()
        else:
            agg_df = plot_df.groupby(group_col)[metric_col].count().reset_index()
        val_col = metric_col
    else:
        plot_df[group_col] = plot_df[group_col].astype(str).str.strip().str.upper()
        agg_df = plot_df[group_col].value_counts().reset_index()
        agg_df.columns = [group_col, 'Count']
        val_col = 'Count'
        
    agg_df = agg_df.sort_values(by=val_col, ascending=False).head(top_n)
    color_palette = ['#22c55e', '#38bdf8', '#a78bfa', '#fbbf24', '#f43f5e', '#34d399', '#60a5fa', '#c084fc']
    
    if chart_type in ["pie", "donut"]:
        hole_val = 0.55 if chart_type == "donut" else 0.0
        fig = px.pie(
            agg_df, 
            names=group_col, 
            values=val_col, 
            hole=hole_val,
            color_discrete_sequence=color_palette
        )
    elif chart_type == "line":
        fig = px.line(
            agg_df, 
            x=group_col, 
            y=val_col, 
            markers=True,
            color_discrete_sequence=['#38bdf8']
        )
    elif chart_type == "treemap":
        fig = px.treemap(
            agg_df, 
            path=[group_col], 
            values=val_col,
            color_discrete_sequence=color_palette
        )
    else:
        fig = px.bar(
            agg_df, 
            x=group_col, 
            y=val_col, 
            text=val_col,
            color=group_col,
            color_discrete_sequence=color_palette
        )
        fig.update_traces(textposition='outside')
        
    fig.update_layout(
        title=dict(text=title, font=dict(family='Space Grotesk', size=16, color='#ffffff')),
        paper_bgcolor='rgba(17, 24, 39, 0.75)',
        plot_bgcolor='rgba(17, 24, 39, 0)',
        font=dict(color='#e2e8f0', family='Plus Jakarta Sans'),
        showlegend=(chart_type in ["pie", "donut"]),
        margin=dict(l=20, r=20, t=50, b=30),
        xaxis=dict(showgrid=False, color='#94a3b8'),
        yaxis=dict(gridcolor='rgba(255, 255, 255, 0.08)', color='#94a3b8')
    )
    return fig

# Pop-up Document Lightbox Modal
@st.dialog("📄 Document Preview", width="large")
def show_preview_modal(file_name, file_bytes, mime_type):
    st.caption(f"Viewing: **{file_name}**")
    if "image" in mime_type:
        img = Image.open(io.BytesIO(file_bytes))
        st.image(img, use_container_width=True)
    elif "pdf" in mime_type:
        st.info(f"📑 PDF File: **{file_name}** ({len(file_bytes)/1024:.1f} KB)")

# Document Upload Section with Multiple File Support
uploaded_files = st.file_uploader(
    "📥 Drop your PDF document(s) or images here", 
    type=["pdf", "png", "jpg", "jpeg"], 
    accept_multiple_files=True
)

if uploaded_files:
    if not api_key:
        st.error("⚠️ System Error: GEMINI_API_KEY not found in Streamlit Secrets. Please add it to your app settings.")
    else:
        files_data = []
        for file in uploaded_files:
            file_bytes = file.read()
            files_data.append((file_bytes, file.type))
        
        col_prev, col_action = st.columns([1.1, 1.9])
        with col_prev:
            st.subheader(f"📄 Uploaded Files ({len(uploaded_files)})")
            for idx, file in enumerate(uploaded_files):
                c_info, c_btn = st.columns([3, 1.2])
                with c_info:
                    icon = "🖼️" if "image" in file.type else "📑"
                    short_name = file.name if len(file.name) < 22 else f"{file.name[:18]}...{file.name[-4:]}"
                    st.markdown(f"""
                    <div style="background: rgba(17, 24, 39, 0.85); border: 1px solid rgba(255, 255, 255, 0.08); border-radius: 8px; padding: 7px 10px; margin-bottom: 6px;">
                        <span style="font-size: 0.85rem; font-weight: 600; color: #e2e8f0;">{icon} {short_name}</span>
                        <div style="font-size: 0.72rem; color: #94a3b8;">{len(files_data[idx][0])/1024:.1f} KB</div>
                    </div>
                    """, unsafe_allow_html=True)
                with c_btn:
                    if st.button("👁️ View", key=f"preview_btn_{idx}", use_container_width=True):
                        show_preview_modal(file.name, files_data[idx][0], file.type)
                
        with col_action:
            st.subheader("⚡ Convert to Excel & Dashboard")
            st.caption(f"Extract and compile all {len(uploaded_files)} file(s) into a unified Excel spreadsheet and visual analytics dashboard.")
            
            extract_clicked = st.button("🚀 Extract Tables & Convert to Excel", type="primary", use_container_width=True)
            loader_container = st.empty()
            
            if extract_clicked:
                loader_html = """
                <div class="glass-loading-card">
                    <div class="spinner-radar-ring"></div>
                    <div class="glass-loading-title">AI Vision Processing & Formatting</div>
                    <div class="glass-loading-desc">
                        Analyzing visual matrix, isolating tabular rows, sanitizing values & building visual dashboard metrics.
                    </div>
                    <div class="status-pills-row">
                        <span class="status-pill">🔍 OCR Matrix Scan</span>
                        <span class="status-pill">🧹 Noise Sanitization</span>
                        <span class="status-pill">📑 Table Structuring</span>
                        <span class="status-pill">📊 .xlsx Synthesis</span>
                    </div>
                </div>
                """
                loader_container.markdown(loader_html, unsafe_allow_html=True)
                
                try:
                    raw_json_str, used_model = execute_extraction_cascade(files_data, api_key)
                    data = json.loads(raw_json_str)
                    st.session_state["extracted_data"] = data
                    st.session_state["model_used"] = used_model
                    loader_container.empty()
                    st.toast(f"Extracted successfully via {used_model}!", icon="⚡")
                except Exception as e:
                    loader_container.empty()
                    st.error(f"Processing Error: {str(e)}")

# =========================================================================
# MAIN RESULTS & 3-MODE WORKSPACE
# =========================================================================
if "extracted_data" in st.session_state:
    data = st.session_state["extracted_data"]
    
    st.markdown("---")
    st.subheader("💡 Executive Summary & Business Insights")
    summary_text = data.get("analysis", "No summary provided.")
    with st.container():
        st.markdown(summary_text)
    
    tables = data.get("tables", [])
    if not tables:
        st.warning("No tables found in the uploaded file(s).")
    else:
        # =========================================================================
        # MODE 1: BASE DATA REVIEW, EDIT & DOWNLOAD
        # =========================================================================
        st.markdown("---")
        st.markdown("""
        <div style="display:flex; align-items:center; gap:12px; margin-top:1.2rem; margin-bottom:0.4rem;">
            <span style="font-size:1.5rem; font-weight:800; color:#ffffff; font-family:'Space Grotesk', sans-serif;">📋 Mode 1: Review & Customize Base Extracted Data</span>
            <span style="font-size:0.75rem; font-weight:600; background:rgba(34,197,94,0.15); color:#4ade80; border:1px solid rgba(34,197,94,0.3); padding:3px 10px; border-radius:9999px;">Data Grid</span>
        </div>
        """, unsafe_allow_html=True)
        st.caption("Double-click any cell below to modify values, add new rows (+), search (🔍), or hide columns (👁) before downloading.")
        
        normalized_dfs = {}
        for idx, tbl in enumerate(tables):
            table_name = tbl.get("table_name", f"Table {idx+1}")
            headers = tbl.get("headers", [])
            rows = tbl.get("rows", [])
            
            cleaned_rows = []
            for r in rows:
                cleaned_rows.append([("" if val is None or str(val).strip() in ["None", "NaN", "nan", "null"] else str(val).strip()) for val in r])
                
            df = pd.DataFrame(cleaned_rows, columns=headers if headers else None)
            df.fillna("", inplace=True)
            df = df.rename(columns={c: normalize_column_header(c) for c in df.columns})
            
            badge_html = f'''
            <div style="display:flex; align-items:center; gap:12px; margin-top:1.5rem; margin-bottom:0.6rem;">
                <span style="font-size:1.15rem; font-weight:700; color:#ffffff; font-family:'Space Grotesk', sans-serif;">📊 {table_name}</span>
                <span style="font-size:0.75rem; font-weight:600; background:rgba(34,197,94,0.15); color:#4ade80; border:1px solid rgba(34,197,94,0.3); padding:3px 10px; border-radius:9999px;">{len(df)} Records</span>
            </div>
            '''
            st.markdown(badge_html, unsafe_allow_html=True)
            
            edited_df = st.data_editor(
                df, 
                key=f"editor_{idx}", 
                num_rows="dynamic", 
                use_container_width=True,
                height=min(420, 45 + len(df) * 35)
            )
            normalized_dfs[table_name] = edited_df

        # Consolidated Master Dataframe
        if len(normalized_dfs) > 1:
            try:
                master_df = pd.concat(list(normalized_dfs.values()), ignore_index=True)
                master_df.dropna(how='all', inplace=True)
                master_df.fillna("", inplace=True)
                if 'SL. NO.' in master_df.columns:
                    master_df['SL. NO.'] = [str(i) for i in range(1, len(master_df) + 1)]
            except Exception:
                master_df = list(normalized_dfs.values())[0]
        else:
            master_df = list(normalized_dfs.values())[0]

        # Base Excel Map
        base_sheets_map = {}
        if len(normalized_dfs) > 1:
            base_sheets_map["Master Combined Records"] = master_df
        for name, df in normalized_dfs.items():
            clean_name = "".join(c for c in name if c.isalnum() or c in (' ', '_', '-')).strip()
            base_sheets_map[clean_name[:31]] = df
            
        base_excel_bytes = generate_base_excel_workbook(base_sheets_map)
        
        st.markdown("<div style='height: 6px;'></div>", unsafe_allow_html=True)
        st.download_button(
            label="📥 Download Clean Base Excel Workbook (.xlsx)",
            data=base_excel_bytes,
            file_name="sheetgen_base_data.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True
        )

        # =========================================================================
        # MODE 2 & MODE 3: VISUAL ANALYTICS & DASHBOARDS
        # =========================================================================
        st.markdown("---")
        st.markdown("""
        <div style="display:flex; align-items:center; gap:12px; margin-top:1.5rem; margin-bottom:0.5rem;">
            <span style="font-size:1.65rem; font-weight:800; color:#ffffff; font-family:'Space Grotesk', sans-serif;">📈 AI Visual Analytics & Dashboard Suite</span>
            <span style="font-size:0.8rem; font-weight:600; background:rgba(56,189,248,0.15); color:#38bdf8; border:1px solid rgba(56,189,248,0.3); padding:3px 12px; border-radius:9999px;">Interactive BI</span>
        </div>
        """, unsafe_allow_html=True)
        st.caption("Review auto-generated dashboards or chat with AI to build custom formula-linked spreadsheets.")

        tab_auto, tab_copilot = st.tabs(["⚡ Mode 2: Smart Auto Dashboard", "💬 Mode 3: Talk with AI (Custom Copilot)"])

        # TAB 1: MODE 2 - SMART AUTO DASHBOARD
        with tab_auto:
            kpi_col1, kpi_col2, kpi_col3 = st.columns(3)
            with kpi_col1:
                st.markdown(f"""
                <div class="kpi-stat-card">
                    <div class="kpi-title">Total Records Parsed</div>
                    <div class="kpi-value">{len(master_df)}</div>
                    <div class="kpi-subtitle">Across {len(normalized_dfs)} sheet(s)</div>
                </div>
                """, unsafe_allow_html=True)
                
            with kpi_col2:
                candidate_cols = [c for c in master_df.columns if c not in ['SL. NO.', 'EMPLOYEE NAME', 'PHONE', 'PHONE NUMBER', 'NO.']]
                cat_col = candidate_cols[0] if candidate_cols else master_df.columns[-1]
                clean_cat = master_df[cat_col].astype(str).str.strip().str.upper().replace("", pd.NA).dropna()
                unique_cnt = clean_cat.nunique()
                st.markdown(f"""
                <div class="kpi-stat-card">
                    <div class="kpi-title">Unique {cat_col[:15]}</div>
                    <div class="kpi-value">{unique_cnt}</div>
                    <div class="kpi-subtitle">Distinct classifications</div>
                </div>
                """, unsafe_allow_html=True)

            with kpi_col3:
                sec_cat_col = candidate_cols[1] if len(candidate_cols) > 1 else cat_col
                clean_sec = master_df[sec_cat_col].astype(str).str.strip().str.upper().replace("", pd.NA).dropna()
                top_val = clean_sec.mode()
                top_name = top_val.iloc[0] if not top_val.empty else "N/A"
                st.markdown(f"""
                <div class="kpi-stat-card">
                    <div class="kpi-title">Dominant {sec_cat_col[:14]}</div>
                    <div class="kpi-value" style="font-size:1.35rem;">{str(top_name)[:18]}</div>
                    <div class="kpi-subtitle">Highest frequency category</div>
                </div>
                """, unsafe_allow_html=True)

            # In-Browser Plotly Preview
            c_chart1, c_chart2 = st.columns(2)
            with c_chart1:
                spec1 = {"chart_type": "bar", "title": f"Distribution by {cat_col}", "group_by_col": cat_col, "aggregation": "count", "top_n": 8}
                fig1 = render_plotly_chart(master_df, spec1)
                st.plotly_chart(fig1, use_container_width=True)

            with c_chart2:
                spec2 = {"chart_type": "donut", "title": f"Proportion of {sec_cat_col}", "group_by_col": sec_cat_col, "aggregation": "count", "top_n": 6}
                fig2 = render_plotly_chart(master_df, spec2)
                st.plotly_chart(fig2, use_container_width=True)

            # Mode 2 Live Formula Excel Download
            st.markdown("<br>", unsafe_allow_html=True)
            smart_excel_bytes = generate_smart_dashboard_excel_workbook(base_sheets_map, master_df)
            st.download_button(
                label="📥 Download Excel with Smart Dashboard & Live Formulas (.xlsx)",
                data=smart_excel_bytes,
                file_name="sheetgen_smart_dashboard_workbook.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_smart_dash_btn",
                use_container_width=True
            )

        # TAB 2: MODE 3 - TALK WITH AI (CUSTOM COPILOT)
        with tab_copilot:
            st.markdown("#### 💬 Ask AI to Build Any Custom Visual Chart")
            st.caption("Click a pre-built question to generate instantly, or describe custom requirements below.")
            
            def run_copilot_query(query_text):
                with st.spinner(f"AI Copilot is analyzing and building dashboard for: '{query_text}'..."):
                    ai_spec = ask_ai_for_dashboard_spec(master_df, query_text, api_key)
                    if ai_spec:
                        st.session_state["custom_ai_spec"] = ai_spec
                        st.session_state["active_chart_prompt"] = query_text
                    else:
                        st.warning("⚠️ Could not generate chart specification. Please try again.")

            # Pre-Built Quick Question Action Chips
            st.markdown("<div style='font-size:0.8rem; font-weight:700; color:#38bdf8; margin-bottom:6px;'>💡 Pre-Built Quick Questions (Click to auto-generate):</div>", unsafe_allow_html=True)
            col_q1, col_q2, col_q3 = st.columns(3)
            with col_q1:
                if st.button("📊 Breakdown by Category / Region", key="pq_1", use_container_width=True):
                    run_copilot_query("Generate a bar chart showing distribution across primary categories or regions.")
            with col_q2:
                if st.button("🍩 Top 5 Contributors (Donut Chart)", key="pq_2", use_container_width=True):
                    run_copilot_query("Show a donut chart of top 5 groups or contributors with highest representation.")
            with col_q3:
                if st.button("🔍 Exception & Remarks Audit", key="pq_3", use_container_width=True):
                    run_copilot_query("Analyze remarks, status flags, or key metrics and display a visual breakdown of exceptions.")

            # Freeform Prompt Input
            current_prompt_val = st.session_state.get("active_chart_prompt", "")
            user_chart_prompt = st.text_input(
                "Describe the chart or metrics you want to build:",
                value=current_prompt_val,
                placeholder="e.g. Compare total headcount across regions and show a pie chart of statuses...",
                key="user_chart_prompt_input"
            )

            if st.button("✨ Generate Custom Visual Dashboard", type="primary", use_container_width=True):
                if user_chart_prompt:
                    run_copilot_query(user_chart_prompt)
                else:
                    st.warning("Please enter a description or click a quick question above.")

            # In-Browser Preview & Custom Excel Download
            if "custom_ai_spec" in st.session_state:
                ai_spec = st.session_state["custom_ai_spec"]
                st.markdown("<br>", unsafe_allow_html=True)
                st.markdown("##### 🎯 Custom AI Dashboard Result")
                
                kpis = ai_spec.get("kpi_cards", [])
                if kpis:
                    kpi_cols = st.columns(min(len(kpis), 3))
                    for i, kpi in enumerate(kpis[:3]):
                        with kpi_cols[i]:
                            st.markdown(f"""
                            <div class="kpi-stat-card">
                                <div class="kpi-title">{kpi.get('label', 'Metric')}</div>
                                <div class="kpi-value" style="font-size:1.5rem;">{kpi.get('value', '-')}</div>
                                <div class="kpi-subtitle">{kpi.get('subtitle', '')}</div>
                            </div>
                            """, unsafe_allow_html=True)

                charts = ai_spec.get("charts", [])
                if charts:
                    c_cols = st.columns(len(charts)) if len(charts) == 2 else [st.container()] * len(charts)
                    for i, c_spec in enumerate(charts):
                        target_col = c_cols[i] if len(charts) == 2 else c_cols[i]
                        with target_col:
                            fig = render_plotly_chart(master_df, c_spec)
                            st.plotly_chart(fig, use_container_width=True)
                            
                if ai_spec.get("ai_insights"):
                    st.info(f"💡 **AI Copilot Insight:** {ai_spec.get('ai_insights')}")
                    
                # Mode 3 Download Button
                custom_ai_excel_bytes = generate_ai_copilot_excel_workbook(base_sheets_map, master_df, ai_spec)
                st.download_button(
                    label="📥 Download AI Customized Dashboard & Live Formulas (.xlsx)",
                    data=custom_ai_excel_bytes,
                    file_name="sheetgen_ai_custom_dashboard.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_ai_dash_btn",
                    use_container_width=True
                )
